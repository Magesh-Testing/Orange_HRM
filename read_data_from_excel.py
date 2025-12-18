import openpyxl
import os

def get_absolute_path(relative_path):
    # Directory where THIS file is located
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_path)

def get_credentials(status_filter, file_path, sheet_name):
    # Convert relative → absolute
    absolute_path = get_absolute_path(file_path)

    if not os.path.exists(absolute_path):
        raise FileNotFoundError(f"Excel file NOT FOUND at: {absolute_path}")

    workbook = openpyxl.load_workbook(absolute_path)
    sheet = workbook[sheet_name]

    credentials = []

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        status = sheet.cell(row=row, column=3).value

        if status == status_filter:
            credentials.append((username, password))

    return credentials

def get_valid_credentials():
    return get_credentials(
        status_filter="Valid",
        file_path="test_data/credentials.xlsx",
        sheet_name="test_data"
    )

def get_invalid_credentials():
    return get_credentials(
        status_filter="Invalid",
        file_path="test_data/credentials.xlsx",
        sheet_name="test_data"
    )